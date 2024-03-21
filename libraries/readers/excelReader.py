import openpyxl
# from variables import paths
# filePath1 = paths.users_test_csv


def read_excel_file(fileName, row_value, column_value):
    wb_obj = openpyxl.load_workbook(fileName, data_only=True)
    sheet_names = wb_obj.get_sheet_names()
    sheet_obj = wb_obj.get_sheet_by_name(sheet_names[0])
    column = int(column_value) + 1  # Here we added '1' because the first column starts from 1 not 0
    m_row = sheet_obj.max_row
    my_list = []  # created an empty list
    for i in range(1,
                   m_row):  # Here I have started the loop from 2 as I want to skip the column heading value in output
        cell_obj = sheet_obj.cell(row=i, column=column)
        print(cell_obj.value)
        my_list.append(cell_obj.value)
    return my_list[int(row_value)]
